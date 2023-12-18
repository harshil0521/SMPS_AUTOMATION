import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import XLUtility
import openpyxl

driver = webdriver.Chrome()
driver.get("https://www.saucedemo.com/")

file = "C:/Users/Harshil.Solanki/Desktop/Files/newusers.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["data"]

print("Total rows:", sheet.max_row)
print("Total Columns:", sheet.max_column)

rows = XLUtility.getRowCount(file, "data") 

for r in range(2, rows+1):
    """Reading credentials"""
    userName = XLUtility.readData(file, "data", r, 1)
    password = XLUtility.readData(file, "data", r, 2)

    """Entering credentials"""
    user_name = driver.find_element(By.NAME, "user-name")
    user_name.clear()
    user_name.send_keys(userName)

    psw = driver.find_element(By.NAME, "password")
    psw.clear()
    psw.send_keys(password)

    login = driver.find_element(By.ID, "login-button")
    login.click()

    if (driver.current_url == "https://www.saucedemo.com/inventory.html"):
        XLUtility.writeData(file, "data", r, 4, "Correct")
        driver.back()
    else:
        driver.refresh()
        XLUtility.writeData(file, "data", r, 4, "Incorrect")
 