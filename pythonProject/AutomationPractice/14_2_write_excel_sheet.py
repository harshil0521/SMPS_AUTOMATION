import openpyxl

"--------------"
"""Same data"""
"--------------"

file = "C:/Users/Prerna Singh/Desktop/Files/Book1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet4"]
# sheet = workbook.active     # if there is only one sheet in workbook

for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(r, c).value = "Hello"

workbook.save(file)


"------------------"
"""Multiple data"""
"------------------"

file = "C:/Users/Prerna Singh/Desktop/Files/Book1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet5"]
# sheet = workbook.active     # if there is only one sheet in workbook

sheet.cell(1, 1).value = 123
sheet.cell(1, 2).value = "ravi"
sheet.cell(1, 3).value = "male"

sheet.cell(2, 1).value = 456
sheet.cell(2, 2).value = "savan"
sheet.cell(2, 3).value = "male"

sheet.cell(3, 1).value = 789
sheet.cell(3, 2).value = "naiya"
sheet.cell(3, 3).value = "female"

workbook.save(file)

