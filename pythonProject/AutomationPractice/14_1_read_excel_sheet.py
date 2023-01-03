"------------------------"
"""Data driven testing"""
"------------------------"

# 1. How to read data from Excel        data = sheet.cell(r,c).value
# 2. How to write data into Excel       sheet.cell(r,c).value = "any msg/data"
# 3. Data driven test case

import openpyxl

file = "C:/Users/Prerna Singh/Desktop/Files/bugsheettest.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row  # Count number of rows in Excel
cols = sheet.max_column  # Count number of columns in Excel

# Reading all the rows & columns from sheet
for r in range(1, rows + 1):
    for c in range(1, cols + 1):
        print(sheet.cell(r, c).value, end="     ")
    print()
