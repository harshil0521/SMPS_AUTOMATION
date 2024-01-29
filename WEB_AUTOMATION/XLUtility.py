import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowNum, columnNum).value


def writeData(file, sheetName, rowNum, columnNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    green_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = green_fill
    workbook.save(file)


def fillRedColor(file, sheetName, rowNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    red_fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = red_fill
    workbook.save(file)
